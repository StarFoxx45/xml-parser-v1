from openpyxl import Workbook, load_workbook

def write_excel(file_name, range_name,col_names, value):
        workbook = Workbook()
        sheet = workbook.active
        for data in value:
            results = data['id']+2
            sheet[range_name+str(results)] = str(data['id'])+", "+str(data['value'])

        for col in col_names:
            
            sheet[str(col['key'])+str(col['col'])]= str(col['col_name'])

        workbook.save(filename=file_name+".xlsx")

def read_excel(file_name, range_name):
        workbook = load_workbook(filename=file_name+".xlsx")
        sheet = workbook.active
        
        value = sheet[range_name]
       

        print(value)

def update_excel(file_name, range_name,col_names, value):
        workbook = load_workbook(filename=file_name+".xlsx")
        sheet = workbook.active
      
        for data in value:
            results = data['id']+2
            sheet[range_name+str(results)] = str(data['id'])+", "+str(data['value'])

        for col in col_names:
            
            sheet[str(col['key'])+str(col['col'])]= str(col['col_name'])

        workbook.save(filename=file_name+".xlsx")

def delete_excel(file_name, range_name):
        workbook = load_workbook(filename=file_name+".xlsx")
        sheet = workbook.active
        
        sheet[range_name] = ""
       
        results = sheet[range_name]

        print([results])
        print("range has been deleted")
